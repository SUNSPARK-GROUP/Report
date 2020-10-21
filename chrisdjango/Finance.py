'''
def hello(request):
    return HttpResponse("Hello world ! ")'''
from django.http import HttpResponse
from django.shortcuts import render,HttpResponseRedirect
from datetime import date
from datetime import timedelta
import cx_Oracle
from siteapp.views import gettotaldata
from siteapp.views import getodate
from siteapp.views import mainmenu
from siteapp.views import submenu
from graphos.sources.simple import SimpleDataSource
from graphos.sources.model import ModelDataSource
from graphos.renderers.gchart import ColumnChart
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import win32com.client as win32
import pymysql
import os
import pyodbc
from operator import itemgetter, attrgetter
import time
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8' 
import datetime

depts=[]
accls=[]
def CONMSSQL214(sqlstr):
    connection214 = pyodbc.connect('DRIVER={SQL Server};SERVER=192.168.0.214;DATABASE=ERPS;UID=apuser;PWD=0920799339')
    cursor = connection214.cursor()
    cursor.execute(sqlstr)
    SQLSTRS = sqlstr[0:6].upper()
    if SQLSTRS == "SELECT":
        TotalSession = cursor.fetchall()
        return TotalSession
        cursor.close()
    else:
        connection214.commit()
def CONMYSQL(sqlstr):
 #  f=open(r'C:\Users\Administrator\Desktop\txtlog\CONMYSQL.txt','w')
  db = pymysql.connect(host='192.168.0.210', port=3306, user='apuser', passwd='0920799339', db='main_eipplus_standard',charset='utf8')
  
  cursor = db.cursor()
  cursor.execute(sqlstr)	

  result = cursor.fetchall()
  urls = [row[0] for row in result]
  return result
  #  f.close()
def CONORACLE(SqlStr):
  # cf = open(r'C:\Users\Administrator\Desktop\txtlog\CONORACLE.txt','w')

  hostname='192.168.0.230'
  sid='E910'
  username='PRODDTA'
  password='E910Jde'
  port='1521'
  dsn = cx_Oracle.makedsn(hostname, port, sid)
  conn = cx_Oracle.connect(username+'/'+password+'@' + dsn,encoding='UTF-8')
  cursor = conn.cursor()

  SQLSTRS = SqlStr[0:6].upper()
  
  if SQLSTRS=="SELECT":
    cursor.execute(SqlStr)
    TotalSession = cursor.fetchall()
    return TotalSession
    cursor.close()
  else:
    # cursor.execute(SqlStr)
    # conn.commit()
    try:
      cursor.execute(SqlStr)
      cursor.close() 
      conn.commit()
    except Exception as e:  
      conn.rollback()   
      raise e          
    finally:     
      conn.close()
  # cf.close()
def showday(wd,sp,dt):#wd->0 today,wd->1 yesterday,wd->-1 tomorrow sp->Divider dt datetype
  t1 = 0-wd
  d=date.today()-timedelta(t1)
  if dt==1 :
    return d.strftime('%d'+sp+'%m'+sp+'%Y')#%Y->2015 %y->15
  else :
    return d.strftime('%Y'+sp+'%m'+sp+'%d')#%Y->2015 %y->15
def company(request):
  context= {}
  comno=[]
  #f=open('C:\\Users\\chris\\chrisdjango\\comerror.txt','w')
  try:
    cstr=request.GET['cstr']
    #f.write(cstr+'\n')
    serdata=CONORACLE("select aban8,abalph from f0101 where (length(aban8)=6 or length(aban8)=7) and abalph like '%"+cstr+"%'  order by aban8")
  except:
    serdata=CONORACLE('select aban8,abalph from f0101 where (length(aban8)=6 or length(aban8)=7) order by aban8')
  for data in serdata:
    tcomno=[]
    tcomno.append(str(data[0]))
    tcomno.append(str(data[1]))
    comno.append(tcomno)
  context['comdata']=comno
  return render(request, 'company.html' ,context)#傳入參數
  #f.close()
def client(request):
  context= {}
  comno=[]
  serdata=CONORACLE("select aban8,abalph from f0101 where ABAT1 = 'C' order by aban8")
  for data in serdata:
    tcomno=[]
    tcomno.append(str(data[0]))
    tcomno.append(str(data[1]))
    comno.append(tcomno)
  context['comdata']=comno
  return render(request, 'client.html' ,context)#傳入參數
def F43121item(request):
  F43121_1=[]
  context= {}
  typels=[{'type':'9','tname':'全部'},{'type':'1','tname':'未付'},{'type':'2','tname':'已付'}]
#   f = open(r'C:\Users\Administrator\Desktop\txtlog\F43121item.txt','w')
  try:
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    sitm=request.GET['sitm']
    eitm=request.GET['eitm']
    tno=request.GET['tno']
    context['tno'] = tno
    compno=request.GET['compno']
    if 	compno=='':
      scompno=" like '%' "
    else:
      scompno= "='"+compno+"' "
    context['Sday'] = sday
    context['Eday'] = eday
    context['compno'] = compno
    context['sitm'] = sitm
    context['eitm'] = eitm
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    #f.write('yy')
    if len(sitm)==0:
      sitemq1=""
      sitemq2=""
    else:
      sitemq1="and f43.prlitm>='"+sitm+"' "
      sitemq2="and f431.pdlitm>='"+sitm+"' "	
    if len(eitm)==0:
      eitemq1=""
      eitemq2=""
    else:
      eitemq1="and f43.prlitm<='"+eitm+"' "
      eitemq2="and f431.pdlitm<='"+eitm+"' "
    '''try:
      ck1=request.GET['Checkbox1']
      context['ck'] ='ck1'
    except:
      ck1='off'
    try:
      ck2=request.GET['Checkbox2']
      context['ck'] ='ck2'
    except:
      ck2='off' '''
    ck1='on' 
    if ck1=='on':
      '''
      title=['名次','地址號','廠商代碼','名稱','採購單號','進貨單號','總帳日期','未結金額','未稅金額','稅金','含稅金額']
      serdata=CONORACLE("select f43.pran8,f01.abalky,f01.abalph,f43.prdoco,f43.prdoc,f43.prdgl,f43.praopn,f43.prarec from f43121 f43,f0101 f01 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'"
                        +"AND ( prmatc=1 OR prdcto='O5') and f01.aban8=f43.pran8  ORDER BY prarec desc ")
      '''  
      title=['<th style="width:6%;">採購單號</th>','<th style="width:6%;">進貨單號</th>','<th style="width:7%;">總帳日期</th>','<th style="width:15%;">廠商名稱</th>','<th style="width:5%;">應付總帳</th>','<th style="width:5%;">料號總帳</th>','<th style="width:5%;">料號</th>','<th style="width:13%;">商品名稱</th>','<th style="width:11%;">商品說明1</th>','<th style="width:6%;">商品說明2</th>','<th style="width:3%;">數量</th>','<th style="width:3%;">單價</th>','<th style="width:4%;">未結金額</th>','<th style="width:4%;">未稅金額</th>','<th style="width:3%;">稅金</th>','<th style="width:4%;">含稅金額</th>']
      uno=format(request.COOKIES['userno'])
      wb = Workbook()	
      ws1 = wb.active	
      ws1.title = "data"
      ws1.append(['森邦(股)應付帳款明細('+sday+'~'+eday+')'])
      ws1.append(['採購單號','進貨單號','總帳日期','地址號','廠商名稱','應付總帳','料號總帳','料號','商品名稱','商品說明1','商品說明2','數量','單價','未結金額','未稅金額','稅金','含稅金額'])
      '''
      serdata=CONORACLE("select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prdgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prdgl,4,3)-1,'yyyymmdd'))) AS prdgl"
                        +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000,f41.imuom1,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
                        +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE 0 END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'  and f43.pran8 like '"+compno+"%' "
                         +"AND ( prmatc=1 OR prdcto='O5') and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm  "#
						 +sitemq1+eitemq1+" union all "
                        +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
                        +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,f431.pddsc1,f431.pddsc2,f431.pdUORG/10000,f41.imuom1,(f431.pdaopn) praopn,(f431.pdaopn) prarec "
                        +",case WHEN pdtxa1='VT' THEN round((f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdaopn)+round((f431.pdaopn)*0.05)  ELSE 0 END otatal,f431.PDlnid,F431.pdan8 "
						+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"' "
                        +"and F431.pdan8 like '"+compno+"%'  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
						 +sitemq2+eitemq2+" ORDER BY pran8,prdgl,prdoc,prlnid  ")
      '''
      #f.write(str(tno)+'\n')
      if tno=='2' :
        #case when f43.prqtypy/10000>0 then round(f43.prarec/(f43.prqtypy/10000),2) else 0 end as price
		
        # f.write(str("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,f431.pddsc1,f431.pddsc2,prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	    #                 +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
        #                 +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
        #                 #+",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'  and f43.pran8 = '"+compno+"' "
        #                 +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8 "+scompno+" "
        #                 +" AND ( prmatc=2 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
		# 				#+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and prmatc=2 ))"				
		# 				+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"'  and x.prmatc=2 AND x.prdcto= f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco = f43.prdoco and x.prlnid = f43.prlnid and x.pran8 = f43.pran8))"
		# 				+" ) a  left join  " #and  f43.praopn<=0
		# 				+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
        #                 +"and F431.pdan8 "+scompno+"  union all "
        #                 #case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
        #                 +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
        #                 +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,f431.pddsc1,f431.pddsc2,f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
        #                 +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
		# 				+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' "
		# 				#+"AND f431.pddoco||f431.pdlnid||f431.pdan8  IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 WHERE prdgl >='"+sd+"'  and prdgl <= '"+ed+"' AND PRDCTO='O5' AND PRDCTO='O5' AND PRMATC=2 )"					
		# 				#+"AND f431.pddoco||f431.pdlnid||f431.pdan8  IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' AND x.PRDCTO='O5' AND x.PRMATC=2 AND x.prdcto=f431.pddcto and x.prdoco = f431.pddoco and x.prlnid = f431.pdlnid and x.pran8= f431.pdan8 )"
		# 				+"AND f431.pddoco||f431.pdlnid||f431.pdan8 IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE  x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' and x.prdoco = f431.pddoco and x.prlnid=f431.pdlnid and x.pran8 = f431.pdan8 )"
		# 				+"AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"'  "
        #                 +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
		# 				 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid  "))
	
        		
        serdata=CONORACLE("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	                    +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
                        +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
                        #+",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'  and f43.pran8 = '"+compno+"' "
                        +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8 "+scompno+" "
                        +" AND ( prmatc=2 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
						#+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and prmatc=2 ))"				
						+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"'  and x.prmatc=2 AND x.prdcto= f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco = f43.prdoco and x.prlnid = f43.prlnid and x.pran8 = f43.pran8))"
						+" ) a  left join  " #and  f43.praopn<=0
						+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
                        +"and F431.pdan8 "+scompno+"  union all "
                        #case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
                        +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
                        +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
                        +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
						+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' "
						#+"AND f431.pddoco||f431.pdlnid||f431.pdan8  IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 WHERE prdgl >='"+sd+"'  and prdgl <= '"+ed+"' AND PRDCTO='O5' AND PRDCTO='O5' AND PRMATC=2 )"					
						#+"AND f431.pddoco||f431.pdlnid||f431.pdan8  IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' AND x.PRDCTO='O5' AND x.PRMATC=2 AND x.prdcto=f431.pddcto and x.prdoco = f431.pddoco and x.prlnid = f431.pdlnid and x.pran8= f431.pdan8 )"
						+"AND f431.pddoco||f431.pdlnid||f431.pdan8 IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE  x.prdgl >='"+sd+"' and x.prdgl <= '"+ed+"' and x.prdoco = f431.pddoco and x.prlnid=f431.pdlnid and x.pran8 = f431.pdan8 )"
						+"AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"'  "
                        +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
						 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid  ")#20190815 f4311 未稅金額 (f431.pdaopn) prarec 改成 f431.pdarec
      if tno=='1' :
        #case when f43.prqtypy/10000>0 then round(f43.prarec/(f43.prqtypy/10000),2) else 0 end as price
        
        # f.write(str("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,f431.pddsc1,f431.pddsc2,prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	    #                 +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
        #                 +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
        #                 +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8  "+scompno+" "
        #                 +"AND ((prprrc = 0) OR (PRAOPN <>0 AND prprrc <>0))  AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
		# 				#+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"' and prmatc=2  ))"
		# 				+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"' and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco = f43.prdoco and x.prlnid = f43.prlnid and x.pran8 = f43.pran8  ))"
		# 				+" ) a  left join  " #20200409取消or f43.praopn>0
		# 				+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
        #                 +"and F431.pdan8 "+scompno+"  union all "
        #                 +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
        #                 #case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
        #                 +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,f431.pddsc1,f431.pddsc2,f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
        #                 +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
		# 				+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' "
		# 				#+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 WHERE prdgl >='"+sd+"'  and prdgl <= '"+ed+"' )"						
		# 				+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' and x.prdoco = f431.pddoco and x.prlnid=f431.pdlnid and x.pran8 = f431.pdan8 )"
		# 				+" AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"' "
        #                 +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
		# 				 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid  "))
		
        serdata=CONORACLE("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	                    +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
                        +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
                        +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8  "+scompno+" "
                        +"AND ((prprrc = 0) OR (PRAOPN <>0 AND prprrc <>0))  AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
						#+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"' and prmatc=2  ))"
						+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"' and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco = f43.prdoco and x.prlnid = f43.prlnid and x.pran8 = f43.pran8  ))"
						+" ) a  left join  " #20200409取消or f43.praopn>0
						+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
                        +"and F431.pdan8 "+scompno+"  union all "
                        +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
                        #case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
                        +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
                        +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
						+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' "
						#+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 WHERE prdgl >='"+sd+"'  and prdgl <= '"+ed+"' )"						
						+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' and x.prdoco = f431.pddoco and x.prlnid=f431.pdlnid and x.pran8 = f431.pdan8 )"
						+" AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"' "
                        +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
						 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid  ")#20190815 f4311 未稅金額 (f431.pdaopn) prarec 改成 f431.pdarec
      if tno=='9' :#case when f43.prqtypy/10000>0 then round(f43.prarec/(f43.prqtypy/10000),2) else 0 end as price
	    
        # f.write(str("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,f431.pddsc1,f431.pddsc2,prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	    #                 +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
        #                 +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
        #                 +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8 "+scompno+" "
        #                 +" AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
		# 				#+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"' and prmatc=2  ))"
		# 				+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"' and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco=f43.prdoco and x.prlnid=f43.prlnid and x.pran8=f43.pran8 ))"
		# 				+" ) a  left join  "#or f43.praopn>0
		# 				+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
        #                 +"and F431.pdan8 "+scompno+" union all "
        
        #                 +"select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,f431.pddsc1,f431.pddsc2,prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
        #                 +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
        #                 #case when f43.prqtypy/10000>0 then round(f43.prarec/(f43.prqtypy/10000),2) else 0 end as price
        #                 +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
        #                 +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8  "+scompno+" "
        #                 +" AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
		# 				#+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and prmatc=2 ))"
		# 				+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"'  and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco=f43.prdoco and x.prlnid=f43.prlnid and x.pran8=f43.pran8 ))"
		# 				+" ) a  left join  "# and  f43.praopn<=0
		# 				+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
        #                 +"and F431.pdan8 "+scompno+" union all "#case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
        
        #                 +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
        #                 +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,f431.pddsc1,f431.pddsc2,f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
        #                 +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
		# 				+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' " 
		# 				#+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 WHERE pddgl >='"+sd+"'  and pddgl <= '"+ed+"' AND PRDCTO='O5' )"
		# 				+"AND f431.pddoco||f431.pdlnid||f431.pdan8 NOT IN (SELECT PRDOCO||PRLNID||PRAN8 FROM F43121 x WHERE x.prdgl >='"+sd+"'  and x.prdgl <= '"+ed+"' AND x.PRDCTO='O5' and x.prdoco = f431.pddoco and x.prlnid = f431.pdlnid and x.pran8 = f431.pdan8 and x.prdcto = f431.pddcto )"
		# 				+"AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"' "
        #                 +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
 		# 				 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid "))
		
        serdata=CONORACLE("select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
	                    +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
                        +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
                        +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8 "+scompno+" "
                        +" AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
						#+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"' and prmatc=2  ))"
						+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"' and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco=f43.prdoco and x.prlnid=f43.prlnid and x.pran8=f43.pran8 ))"
						+" ) a  left join  "#or f43.praopn>0
						+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
                        +"and F431.pdan8 "+scompno+" union all "
        
                        +"select prdoco,prdoc,prrcdj,abalph,prglc,imglpt,imlitm,imsrtx,Replace (f431.pddsc2,'=',''),Replace (f431.pddsc2,'=',''),prqtypy,price,praopn,prarec,tax,otatal ,prlnid,pran8 from "
                        +"(select f43.prdoco,f43.prdoc,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.prrcdj,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.prrcdj,4,3)-1,'yyyymmdd'))) AS prrcdj"
                        #case when f43.prqtypy/10000>0 then round(f43.prarec/(f43.prqtypy/10000),2) else 0 end as price
                        +",f01.abalph,f43.prglc,f41.imglpt,f41.imlitm,f41.imsrtx,f41.imdsc1,f41.imdsc2,f43.prqtypy/10000 as prqtypy ,round(f43.prprrc/10000,4) as price,(f43.praopn) praopn,(f43.prarec) prarec,case WHEN prtxa1='VT' THEN round((f43.prarec)*0.05) ELSE 0 END tax "
                        +",case WHEN prtxa1='VT' THEN (f43.prarec)+round((f43.prarec)*0.05)  ELSE (f43.prarec) END otatal,f43.prlnid,f43.pran8 from f43121 f43,f0101 f01,f4101 f41 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and f43.pran8  "+scompno+" "
                        +" AND ( prmatc=1 ) and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm "+sitemq1+eitemq1
						#+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 where prrcdj >='"+sd+"' and prrcdj <= '"+ed+"'  and prmatc=2 ))"
						+" and (f43.prdoco||f43.prlnid||f43.pran8  in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"'  and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco=f43.prdoco and x.prlnid=f43.prlnid and x.pran8=f43.pran8 ))"
						+" ) a  left join  "# and  f43.praopn<=0
						+"f4311 f431 on f431.pddoco=a.prdoco and F431.pdlitm=a.imlitm and f431.pdlnid=a.prlnid  "
                        +"and F431.pdan8 "+scompno+" union all "#case WHEN f431.pdUORG/10000 >0 then round(f431.pdaopn/(f431.pdUORG/10000),2) else 0 end as price
        
                        +"select  f431.pddoco,f431.pdDOCC,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f431.pddgl,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f431.pddgl,4,3)-1,'yyyymmdd'))) "
                        +"AS prdgl,f01.abalph,f431.pdglc,f41.imglpt,f41.imlitm,f41.imsrtx,Replace (f431.pddsc1,'=',''),Replace (f431.pddsc2,'=',''),f431.pdUORG/10000,round(f431.pdprrc/10000,4) as price,f431.pdaopn,f431.pdarec+f431.pdaopn "
                        +",case WHEN pdtxa1='VT' THEN round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE 0 END  tax ,case WHEN pdtxa1='VT' THEN (f431.pdarec+f431.pdaopn)+round((f431.pdarec)*0.05+(f431.pdaopn)*0.05)  ELSE (f431.pdarec+f431.pdaopn) END otatal,f431.PDlnid,F431.pdan8 "
						+" from f4311 f431,f0101 f01,f4101 f41 where f431.pddcto='O5' " 
						#+" and (f43.prdoco||f43.prlnid||f43.pran8 not in (select prdoco||prlnid||pran8 from f43121 x where x.prrcdj >='"+sd+"' and x.prrcdj <= '"+ed+"' and x.prmatc=2 AND x.prdcto=f43.prdcto AND x.prmatc=f43.prmatc and x.prdoco=f43.prdoco and x.prlnid=f43.prlnid and x.pran8=f43.pran8 ))"
						+"AND F431.pddgl >='"+sd+"'  and F431.pddgl <= '"+ed+"' "
                        +"and F431.pdan8 "+scompno+"  and f01.aban8=F431.pdan8 and f41.imlitm=F431.pdlitm "
 						 +sitemq2+eitemq2+" ORDER BY pran8,prrcdj,prdoc,prlnid  ")  #20190815 f4311 未稅金額 (f431.pdaopn) prarec 改成 f431.pdarec f431.pdaopn 改成F431.pdaexp
        

      context['reportmes']='<B>森邦(股)應付帳款明細('+sday+'~'+eday+')</B>'
      context['title'] = title
      ds=0
      untax=0
      tax=0
      sum=0
      checkcpt=0
      tuntax=0#未稅金額小計
      ttax=0#稅金小計
      tsum=0#含稅金額小計
      tan8=''
      cellf=PatternFill(start_color="AACF91", end_color="AACF91", fill_type="solid")
      maxc=2
      for data in serdata:
        tf43=[]
        maxc=maxc+1
        if tan8!='' and tan8!=str(data[17]):
          ws1.append(['','','','','','','','','','','','','','小計',format(tuntax,','),format(ttax,','),format(tsum,',')])	
          ws1.cell(row=maxc, column=14).fill=cellf	
          ws1.cell(row=maxc, column=15).fill=cellf	
          ws1.cell(row=maxc, column=16).fill=cellf	
          ws1.cell(row=maxc, column=17).fill=cellf	
          tuntax=0#未稅金額小計
          ttax=0#稅金小計
          tsum=0#含稅金額小計
          maxc=maxc+1		  
        #ds=ds+1
        #tf43.append('<td style="width:4%;">'+str(ds)+'</td>')
        tf43.append('<td style="width:6%;">'+str(data[0])+' '+'</td>')
        if str(data[1])=='0':
          tf43.append('<td style="width:6%;">O5</td>')
        else:
          tf43.append('<td style="width:6%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:7%;">'+str(data[2])+'</td>')
        tf43.append('<td style="width:15%;">'+str(data[3])+'</td>')
        if str(data[4])!=str(data[5]):
          checkcpt=checkcpt+1
        tf43.append('<td style="width:5%;">'+str(data[4])+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[5])+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[6])+'</td>')
        tf43.append('<td style="width:13%;">'+str(data[7])+'</td>')
        tf43.append('<td style="width:11%;">'+str(data[8])+'</td>')
        tf43.append('<td style="width:6%;">'+str(data[9])+'</td>')
        tf43.append('<td style="width:3%;">'+str(data[10])+'</td>')
        tf43.append('<td style="width:3%;">'+str(data[11])+'</td>')
        tf43.append('<td style="width:4%;">'+format(data[12],',')+'</td>')#未結金額
        tf43.append('<td style="width:4%;">'+format(data[13],',')+'</td>')#未稅金額
        untax=untax+int(data[13])        
        tuntax=tuntax+int(data[13])
        tf43.append('<td style="width:3%;">'+format(data[14],',')+'</td>')#稅金
        tax=tax+int(data[14])
        ttax=ttax+int(data[14])
        tf43.append('<td style="width:4%;">'+format(data[15],',')+'</td>')#含稅金額
        sum=sum+int(data[15])
        tsum=tsum+int(data[15])
        F43121_1.append(tf43)
        tan8=str(data[17])
        ws1.append([str(data[0]),str(data[1]),str(data[2]),str(data[17]),str(data[3]),str(data[4]),str(data[5]),str(data[6]),str(data[7]),str(data[8]),str(data[9]),str(data[10]),str(data[11]),str(data[12]),str(data[13]),str(data[14]),str(data[15])])
      ws1.append(['','','','','','','','','','','','','','小計',format(tuntax,','),format(ttax,','),format(tsum,',')])	
      maxc=maxc+1
      ws1.cell(row=maxc, column=14).fill=cellf	
      ws1.cell(row=maxc, column=15).fill=cellf	
      ws1.cell(row=maxc, column=16).fill=cellf	
      ws1.cell(row=maxc, column=17).fill=cellf
      cellf=PatternFill(start_color="FFFF37", end_color="FFFF37", fill_type="solid")
      ws1.append(['','','','','','','','','','','','','','總計',format(untax,','),format(tax,','),format(sum,',')])	
      maxc=maxc+1
      ws1.cell(row=maxc, column=14).fill=cellf	
      ws1.cell(row=maxc, column=15).fill=cellf	
      ws1.cell(row=maxc, column=16).fill=cellf	
      ws1.cell(row=maxc, column=17).fill=cellf
      F43121_1.append(['<td style="width:6%;"></td>','<td style="width:6%;"></td>','<td style="width:7%;"></td>','<td style="width:15%;"></td>','<td style="width:5%;"></td>','<td style="width:5%;"></td>','<td style="width:5%;"></td>','<td style="width:15%;"></td>','<td style="width:3%;"></td>','<td style="width:3%;"></td>','<th style="width:3%;"></th>','<th style="width:4%;"></th>','<td style="width:8%;">總計</td>','<td style="width:8%;">'+format(untax,',')+'</td>','<td style="width:7%;">'+format(tax,',')+'</td>','<td style="width:8%;">'+format(sum,',')+'</td>'])
      context['F43121_1'] = F43121_1
      fday=showday(0,'',0) #今天日期
      try:
        os.remove('C:\\Users\\Administrator\\chrisdjango\\MEDIA\\'+fday+uno+'應付帳款明細.xlsx')
      except OSError as e:
        print(e)
      wb.save('C:\\Users\\Administrator\\chrisdjango\\MEDIA\\'+fday+uno+'應付帳款明細.xlsx')
      #wb.save('C:\\Users\\chris\\chrisdjango\\MEDIA\\'+fday+uno+'應付帳款明細.xlsx')
      context['typels'] = typels
      context['efilename']=fday+uno+'應付帳款明細.xlsx'
      #f.write(str(len(F43121))+'\n')    
    if str(len(F43121_1))=='0':
        context['mess']='查無資料'
    else:
      #f.write("共 "+str(len(F43121))+" 筆資料"+'\n')
      if checkcpt==0:
        context['mess']="共 "+str(len(F43121_1)-1)+" 筆資料"	
      else:
        context['mess']="共 "+str(len(F43121_1)-1)+" 筆資料，有"+str(checkcpt)+"筆總帳號不符"	  
  except:
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
    context['typels'] = typels
  return render(request, 'F43121item.html',context )#傳入參數
  #f.write(str(context))
#   f.close() 
def F43121(request):
  F43121=[]
  context= {}
  #f=open('C:\\Users\\chris\\chrisdjango\\error.txt','w')
  try:
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    compno=request.GET['compno']	
    context['Sday'] = sday
    context['Eday'] = eday
    context['compno'] = compno
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    try:
      ck1=request.GET['Checkbox1']
      context['ck'] ='ck1'
    except:
      ck1='off'
    try:
      ck2=request.GET['Checkbox2']
      context['ck'] ='ck2'
    except:
      ck2='off'
    if ck1=='on':
      #f.write('ck1'+'\n')
      '''
      title=['名次','地址號','廠商代碼','名稱','採購單號','進貨單號','總帳日期','未結金額','未稅金額','稅金','含稅金額']
      serdata=CONORACLE("select f43.pran8,f01.abalky,f01.abalph,f43.prdoco,f43.prdoc,f43.prdgl,f43.praopn,f43.prarec from f43121 f43,f0101 f01 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'"
                        +"AND ( prmatc=1 OR prdcto='O5') and f01.aban8=f43.pran8  ORDER BY prarec desc ")
      '''
      	  
      title=['<th style="width:4%;">名次</th>','<th style="width:5%;">地址號</th>','<th style="width:8%;">廠商代碼</th>','<th style="width:30%;">名稱</th>','<th style="width:10%;">未結金額</th>','<th style="width:10%;">未稅金額</th>','<th style="width:10%;">稅金</th>','<th style="width:10%;">含稅金額</th>']
      '''
      wb = Workbook()	
      ws1 = wb.active	
      ws1.title = "data"
      ws1.append(['名次','地址號','廠商代碼','名稱','採購單號','進貨單號','總帳日期','未結金額','未稅金額','稅金','含稅金額'])
      '''
      serdata=CONORACLE("select f43.pran8,f01.abalky,f01.abalph,sum(f43.praopn) praopn,sum(f43.prarec) prarec,round(sum(f43.prarec)*0.05) tax,sum(f43.prarec)+round(sum(f43.prarec)*0.05) otatal from f43121 f43,f0101 f01 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'"
                        +"AND ( prmatc=1 OR prdcto='O5') and f43.pran8 like '%"+compno+"%' and f01.aban8=f43.pran8 group by f43.pran8,f01.abalky,f01.abalph ORDER BY prarec desc ")
      context['reportmes']='<B>森邦(股) 進貨總金額排名('+sday+'~'+eday+')</B>'
      context['title'] = title
      ds=0
      for data in serdata:
        tf43=[]
        ds=ds+1
        tf43.append('<td style="width:4%;">'+str(ds)+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[0])+' '+'</td>')
        tf43.append('<td style="width:8%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:30%;">'+str(data[2])+'</td>')
        tf43.append('<td style="width:10%;">'+format(data[3],',')+'</td>')#未結金額
        tf43.append('<td style="width:10%;">'+format(data[4],',')+'</td>')#未稅金額
        tf43.append('<td style="width:10%;">'+format(data[5],',')+'</td>')#稅金
        tf43.append('<td style="width:10%;">'+format(data[6],',')+'</td>')#含稅金額
        F43121.append(tf43)
        #f.write(str(tf43))
      context['F43121'] = F43121
      #f.write(str(len(F43121))+'\n')
    if ck2=='on':
      title=['<th style="width:4%;">名次</th>','<th style="width:10%;">地址號</th>','<th style="width:10%;">第二料號</th>','<th style="width:30%;">商品名稱</th>','<th style="width:10%;">未稅金額</th>','<th style="width:10%;">數量</th>']#,'<th style="width:20%;">廠商名稱</th>'
      serdata=CONORACLE("SELECT F43.PRAN8,F43.PRLITM,F41.imdsc1,sum(f43.prarec) as prarec,sum(F43.PRUORG/1000) as qty FROM F43121 F43,F0101 F01 ,F4101 F41 where prdgl >='"+sd+"' and prdgl <= '"+ed+"'"
                        +"AND ( prmatc=1 OR prdcto='O5') and f01.aban8=f43.pran8 and f41.imlitm=f43.prlitm and f43.pran8 like '%"+compno+"%'  group by F43.PRAN8,F43.PRLITM,F41.imdsc1 order by sum(f43.prarec) desc ")#,F01.ABalph
      context['title'] = title
      context['reportmes']='<B>森邦(股) 商品總金額排名('+sday+'~'+eday+')</B>'
      ds=0
      for data in serdata:
        tf43=[]
        ds=ds+1
        tf43.append('<td style="width:4%;">'+str(ds)+'</td>')
        tf43.append('<td style="width:10%;">'+str(data[0])+' '+'</td>')
        #tf43.append('<td style="width:20%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:10%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:30%;">'+str(data[2])+'</td>')
        tf43.append('<td style="width:10%;">'+format(data[3],',')+'</td>')#未稅金額
        tf43.append('<td style="width:10%;">'+format(data[4],',')+'</td>')#數量
        F43121.append(tf43)
      context['F43121'] = F43121	
    if str(len(F43121))=='0':
        context['mess']='查無資料'
    else:
      #f.write("共 "+str(len(F43121))+" 筆資料"+'\n')
      context['mess']="共 "+str(len(F43121))+" 筆資料"		
  except:
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
  return render(request, 'F43121.html',context )#傳入參數
  #f.write(str(context))
  #f.close()    
def F03B11(request):
  context= {}
  try:
    #f=open(r'C:\Users\chris\chrisdjango\errorf03b11.txt','a+')
    F03B11=[]
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    RPDCT=request.GET['RPDCT']
    AN8=request.GET['AN8']
    context['sRPDCT']=RPDCT
    context['Sday'] = sday
    context['Eday'] = eday
    context['sAN8'] = AN8
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    #f.write(request.GET['commit']+'\n')
    #f.write(sday+'-'+eday+'\n')
    try:
      ck1=request.GET['Checkbox1']
      context['ck'] ='ck1'
    except:
      ck1='off'
    noan8=""
    if ck1=='on':
      noan8=" AND RPAN8 not in ("+"'6000009','6000010') "
    #f.write(ck1+'\n')  
    if request.GET['commit']== '查詢發票':  
      evidata=CONORACLE("SELECT a.* ,(a.NEW_RPATXA-a.RPATXA) gap FROM (SELECT RPAN8,TO_CHAR(rpalph) rpalph,SUM(RPAG) RPAG,SUM(RPATXA) RPATXA,ROUND(SUM(RPAG)/1.05,0) NEW_RPATXA"
                  +",SUM(RPSTAM) RPSTAM,SUM(RPAG)-ROUND(SUM(RPAG)/1.05,0) NEW_RPSTAM,round(SUM(RPATAD),0) RPATAD FROM f03b11 "
                  +"WHERE RPDGJ>='"+sd+"' AND RPDGJ<='"+ed+"' AND rpurrf='               ' and RPDCT in('RI','RM')  AND RPDCT like '%"+RPDCT+"%' AND RPAN8 LIKE '"+AN8+"%' "+noan8#and RPAN8='1000548' and RPAN8 like '4%'
                  +"GROUP BY RPAN8,rpalph) a  where (RPATXA-NEW_RPATXA) <> 0 ORDER BY RPAN8") #20190522 and RPDCT<>'RB'  AND rpurrf='               '   
    elif request.GET['commit']== '修正稅額':  
      f=open(r'C:\Users\chris\chrisdjango\errorf03b11.txt','a+')    
      evidata=CONORACLE("SELECT a.* ,(a.NEW_RPATXA-a.RPATXA) gap FROM (SELECT RPAN8,TO_CHAR(rpalph) rpalph,SUM(RPAG) RPAG,SUM(RPATXA) RPATXA,ROUND(SUM(RPAG)/1.05,0) NEW_RPATXA"
                  +",SUM(RPSTAM) RPSTAM,SUM(RPAG)-ROUND(SUM(RPAG)/1.05,0) NEW_RPSTAM,round(SUM(RPATAD),0) RPATAD FROM f03b11 "
                  +"WHERE RPDGJ>='"+sd+"' AND RPDGJ<='"+ed+"' AND rpurrf='               ' and RPDCT in('RI','RM')  AND RPDCT like '%"+RPDCT+"%'  AND RPAN8 LIKE '"+AN8+"%' "+noan8#and RPAN8='1000548' and RPAN8 like '4%'
                  +"GROUP BY RPAN8,rpalph) a  where (RPATXA-NEW_RPATXA) <> 0 ORDER BY RPAN8")#20190522 and RPDCT<>'RB' 
      for e in evidata:
        i=int(e[8])
        evid=CONORACLE("SELECT RPDOC,RPAN8,rpalph,RPAG,RPATXA,RPSTAM, RPATAD FROM f03b11 "
                 +"WHERE RPDGJ>='"+sd+"' AND RPDGJ<='"+ed+"' AND rpurrf='               ' and RPDCT in('RI','RM') and RPAN8='"+str(e[0])+"' "
                 +"order by RPAG DESC")#20190522 and RPDCT<>'RB' 
        for d in evid:
          f.write("update f03b11 set RPATXA='"+str(int(d[4])+i)+"',RPSTAM='"+str(int(d[5])-i)+"',RPATAD='"+str(int(d[4])+i)+"' where RPDOC='"+str(d[0])+"' "+'\n')
          f.write("f03b11原未稅/稅金:"+str(int(d[4]))+"/"+str(int(d[5])-i)+" 修改後:"+str(int(d[4])+i)+"/ "+str(int(d[5])-i)+"\n")
          CONORACLE("update f03b11 set RPATXA='"+str(int(d[4])+i)+"',RPSTAM='"+str(int(d[5])-i)+"',RPATAD='"+str(int(d[4])+i)+"' where RPDOC='"+str(d[0])+"' ")
          f.write("select gldoc,glani,glaa from f0911 where GLDOC = '"+str(d[0])+"' and glani not in('      100000.5101            ','      100000.5201            ','      100000.1201            ',      100000.1202            )  order by abs(glaa) desc"+'\n')
          F0911=CONORACLE("select gldoc,glani,glaa from f0911 where GLDOC = '"+str(d[0])+"' and glani not in('      100000.5101            ','      100000.5201            ','      100000.1201            ',      100000.1202            )  order by abs(glaa) desc")
          for ee in F0911: #str(int(d[4])+i) 調整後未稅金
            aa=int(d[4])+i-ee[2]#差額(f03b11未稅額-f0911未稅額)
            f.write("update f0911 set glaa=glaa-("+str(i)+") where GLDOC='"+str(ee[0])+"' and glani='"+str(ee[1])+"' "+'\n')
            f.write("f0911差額:"+str(i)+"\n")
            updatef0911=CONORACLE("update f0911 set glaa=glaa-("+str(i)+") where GLDOC='"+str(ee[0])+"' and glani='"+str(ee[1])+"' ")        
            break
          break
      #f.close()
      evidata=CONORACLE("SELECT a.* ,(a.NEW_RPATXA-a.RPATXA) gap FROM (SELECT RPAN8,TO_CHAR(rpalph) rpalph,SUM(RPAG) RPAG,SUM(RPATXA) RPATXA,ROUND(SUM(RPAG)/1.05,0) NEW_RPATXA"
                  +",SUM(RPSTAM) RPSTAM,SUM(RPAG)-ROUND(SUM(RPAG)/1.05,0) NEW_RPSTAM,round(SUM(RPATAD),0) RPATAD FROM f03b11 "
                  +"WHERE RPDGJ>='"+sd+"' AND RPDGJ<='"+ed+"' AND rpurrf='               ' and RPDCT in('RI','RM') AND RPDCT like '%"+RPDCT+"%'  AND RPAN8 LIKE '"+AN8+"%' "+noan8#and RPAN8='1000548' and RPAN8 like '4%'
                  +"GROUP BY RPAN8,rpalph) a  where (RPATXA-NEW_RPATXA) <> 0 ORDER BY RPAN8")#20190522 and RPDCT<>'RB' 
    for data in evidata:
      TF03B11=[]
      TF03B11.append(str(data[0]))
      TF03B11.append(str(data[1]))
      TF03B11.append(format(data[2],','))#含稅金額
      TF03B11.append(format(data[3],','))#未稅金額
      TF03B11.append(format(data[5],','))#稅    金
      TF03B11.append(format(data[6],','))#正確稅額
      TF03B11.append(format(data[8],','))#稅 額 差
      F03B11.append(TF03B11)
    context['F03B11']=F03B11
    if len(F03B11)==0:
      context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(F03B11))+" 筆資料"
    context['reportmes']='<B>森邦(股) 發票稅額修正查詢明細('+sday+'~'+eday+')</B>'	
    #f.close()	
  except:
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
    context['ck'] ='ck1'
  return render(request, 'F03B11.html',context )#傳入參數
  
def F0911(request): 
  global depts
  global accls
  context= {}
  F0911=[]
  #f=open(r'C:\Users\chris\chrisdjango\error.txt','w')
  try:    
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    context['Sday'] = sday
    context['Eday'] = eday
    try:
      TOEXCEL=request.GET['2EXCEL']
    except:
      TOEXCEL=''
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    mcu=request.GET['depts']
    macct=request.GET['acct']
    subacct=request.GET['subacct']
    acct=macct.replace(' ','')+'.'+subacct
    context['acct'] = macct
    context['subacct'] = subacct
    context['dept'] = mcu
    lastsum=0
    wb = Workbook()	
    ws1 = wb.active	
    ws1.title = "data"
    ws1.append(['森邦(股)會計科目餘額明細('+mcu+'_'+macct+'_'+subacct+'_'+sday+'~'+eday+')'])
    ws1.append(['總帳日期','科目代碼','說明','其他備註','借方金額','買方金額','科目餘額','批次號碼','地址號'])
    '''
    f.write("select 0+sum(AA)+sum(AA1) from (select case when F09.GLAA>=0 then f09.GLAA ELSE 0 END AS  AA,case when F09.GLAA <0 then f09.GLAA ELSE 0 END AS  AA1 from F0911 F09 "
                      +"  where GLDGJ<'"+sd+"' and GLDGJ>='114001'  and  f09.globj like '%"+macct.replace(' ','')+"%' and f09.glsub like'%"+subacct.replace(' ','')+"%'"
					  +" and glmcu like '%"+mcu+"' AND  GLLT IN ('AA'))"+'\n')
    '''
    if acct[:1] =='1' or acct[:1] =='3' or acct[:1] =='5' or acct[:1] =='6' or acct[:1] =='8':
      
      ORADB = CONORACLE("select 0+sum(AA)+sum(AA1) from (select case when F09.GLAA>=0 then f09.GLAA ELSE 0 END AS  AA,case when F09.GLAA <0 then f09.GLAA ELSE 0 END AS  AA1 from F0911 F09 "
                      +"  where GLDGJ<'"+sd+"' and GLDGJ>='114001'  and  f09.globj like '%"+macct.replace(' ','')+"%' and f09.glsub like'%"+subacct.replace(' ','')+"%'"
					  +" and glmcu like '%"+mcu+"' AND  GLLT IN ('AA'))")
    else:      
      ORADB = CONORACLE("select 0-sum(AA)-sum(AA1) from (select case when F09.GLAA>=0 then f09.GLAA ELSE 0 END AS  AA,case when F09.GLAA <0 then f09.GLAA ELSE 0 END AS  AA1 from F0911 F09 "
                      +"  where GLDGJ<'"+sd+"'  and GLDGJ>='114001'  and  f09.globj like '%"+macct.replace(' ','')+"%' and f09.glsub like'%"+subacct.replace(' ','')+"%'"
					  +" and glmcu like '%"+mcu+"' AND  GLLT IN ('AA'))")
    
    for o in ORADB:
      try:        
        lastsum=int(o[0])
        F0911.append(['',acct,'','','','',format(lastsum,','),'',''])
        #ws1.append(['',acct,'','','','',format(lastsum,','),'',''])
        ws1.append(['',acct,'','','','',int(lastsum),'',''])
      except:        
        #f.write('lastsum:'+int(lastsum)+'\n')
        F0911.append(['',acct,'','','','',format(lastsum,','),'',''])
        #ws1.append(['',acct,'','','','',format(lastsum,','),'',''])
        ws1.append(['',acct,'','','','',int(lastsum),'',''])
    '''
    f.write("select TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(F09.GLDGJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(F09.GLDGJ,4,3)-1,'yyyymmdd'))) AS GLDGJ,F09.GLOBJ||'.'||F09.GLSUB,f01.abalph,F09.GLEXR,case when F09.GLAA>=0 then f09.GLAA ELSE 0 END AS  AA "
                      +",case when F09.GLAA <0 then f09.GLAA ELSE 0 END AS  AA1,'' AS ACCTLAT,F09.GLICU,F09.GLAN8 "
                      +" from F0911 F09 left join f0101 f01 on f01.aban8=f09.glan8 where GLDGJ>='"+sd+"' and GLDGJ<='"+ed+"'  and  f09.globj like '%"+macct.replace(' ','')+"%' and f09.glsub like'%"+subacct.replace(' ','')+"%'"
                      +" and glmcu like '%"+mcu+"' AND  GLLT IN ('AA') order by F09.GLDGJ"+'\n')
    '''
    ORADB = CONORACLE("select TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(F09.GLDGJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(F09.GLDGJ,4,3)-1,'yyyymmdd'))) AS GLDGJ,F09.GLOBJ||'.'||F09.GLSUB,f01.abalph,F09.GLEXR,case when F09.GLAA>=0 then f09.GLAA ELSE 0 END AS  AA "
                      +",case when F09.GLAA <0 then f09.GLAA ELSE 0 END AS  AA1,'' AS ACCTLAT,F09.GLICU,F09.GLAN8 "
                      +" from F0911 F09 left join f0101 f01 on f01.aban8=f09.glan8 where GLDGJ>='"+sd+"' and GLDGJ<='"+ed+"'  and  f09.globj like '%"+macct.replace(' ','')+"%' and f09.glsub like'%"+subacct.replace(' ','')+"%'"
                      +" and glmcu like '%"+mcu+"' AND  GLLT IN ('AA') order by F09.GLDGJ")	#F09.glexa ,F09.GLVINV,'' AS TAX,'' AS TAXP,'' AS INVTYPE
    subtotal1=0 
    subtotal2=0
    for o in ORADB:
      TF0911=[]	  
      TF0911.append(str(o[0]))      
      TF0911.append(str(o[1]))
      TF0911.append(str(o[2]))
      TF0911.append(str(o[3]))
      #TF0911.append(format(o[4],','))#借方
      #TF0911.append(format(o[5],','))#貸方
      TF0911.append(int(o[4]))#借方
      TF0911.append(int(o[5]))#貸方
      if acct[:1] =='1' or acct[:1] =='3' or acct[:1] =='5' or acct[:1] =='6' or acct[:1] =='8':
        lastsum=lastsum+int(o[4])+int(o[5])        
      else:
        lastsum=lastsum-int(o[4])-int(o[5])
        #f.write(acct+'\n')
      subtotal1=subtotal1+int(o[4])
      subtotal2=subtotal2+int(o[5])
      #TF0911.append(format(lastsum,','))#餘額
      TF0911.append(int(lastsum))#餘額
      TF0911.append(str(o[7]))
      TF0911.append(str(o[8]))
      '''
      TF0911.append(str(o[9]))
      TF0911.append(str(o[10]))
      TF0911.append(str(o[11]))
      TF0911.append(str(o[12]))
      '''
      #f.write(str(TF0911))
      F0911.append(TF0911) 
      ws1.append(TF0911)
    context['F0911']=F0911
    context['deptls']=depts
    #context['acctls']=accls
    if len(F0911)>1:
      context['mess']=mcu+'_'+acct+"共 "+str(len(F0911)-1)+" 筆資料"
      F0911.append(['','','','',format(subtotal1,','),format(subtotal2,','),'','',''])
    else:
      context['mess']=mcu+'_'+acct+'查無資料'
    context['reportmes']='<B>森邦(股) 會計科目餘額查詢明細('+sday+'~'+eday+')</B>'
    wb.save('C:\\Users\\Administrator\\chrisdjango\\MEDIA\\'+mcu+'_'+macct+'_'+subacct+'_'+'會計科目餘額明細.xlsx')#正式路徑
    #wb.save('C:\\Users\\chris\\chrisdjango\\MEDIA\\'+mcu+'_'+macct+'_'+subacct+'_'+'會計科目餘額明細.xlsx')#測試路徑
    context['efilename']=mcu+'_'+macct+'_'+subacct+'_'+'會計科目餘額明細.xlsx'
    '''
    if TOEXCEL =='匯出EXCEL' :
      wb = openpyxl.Workbook()
      ws = wb.active
      ws.append(['總帳日期','科目代碼','說明','其他備註','借方金額','買方金額','科目餘額','批次號碼','地址號','發票號碼','課稅','稅率','發票格式'])
      
      for e in range(len(F0911)):
        ws.append(F0911[e])
      
      wb.save(r'c:\create_sample.xlsx')
      excel = win32.gencache.EnsureDispatch('Excel.Application')
      wb = excel.Workbooks.Open(r'c:\create_sample.xlsx')
      excel.Visible = True
    f.write(th75)    
    '''	
  except:
    s='' 
    accls=[]
    depts=[]
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
    context['F0911']=F0911
    context['mess']=''
    context['macct'] =''
    '''
    accts = CONORACLE("select distinct(replace(gmobj,' ','')||'.'||replace(gmsub,' ',''))  from f0901 where GMLDA in (6,7)  order by replace(gmobj,' ','')||'.'||replace(gmsub,' ','') ") 
    for a in accts:
      tal={'mac':str(a[0])}
      accls.append(tal)
    context['acctls']=accls
    '''
    dept = CONORACLE("select distinct(mcmcu) mcmcu,mcmcu||'-'||mcdc from(select (replace(mcmcu,' ' ,'')) mcmcu,mcdc from f0006)  order by mcmcu")
    depts.append({'depno':'','depname':''})
    for a in dept:
      tdp={'depno':str(a[0]),'depname':str(a[1])}
      depts.append(tdp)
    context['deptls']=depts
  #f.close()
  return render(request, 'F0911.html',context )#傳入參數
def F4211(request):
  context= {}
  F4211=[]
  #f=open(r'C:\Users\chris\chrisdjango\error.txt','w')
  try:    
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    context['Sday'] = sday
    context['Eday'] = eday
    stype=''
    ST=[]
    try:
      ck1=request.GET['Checkbox1']
    except:
      ck1='off'
    try:
      ck2=request.GET['Checkbox2']
    except:
      ck2='off'
    if (ck1 == 'on'):
      ST.append('SD')
      context['CK1'] = 'on'
    if ck2 == 'on':
      ST.append('S1')	  
      context['CK2'] = 'on'
    if len(ST)>0:
      stype='AND SDDCTO not in ('
      for i in range(len(ST)):
        if i<len(ST)-1:
          stype=stype+"'"+ST[i]+"',"
        else:
          stype=stype+"'"+ST[i]+"'"
      stype=stype+')'
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    sSDDCTO=request.GET['SDDCTO'].upper()
    if len(sSDDCTO)==0:
      qSDDCTO=''
    else:
      qSDDCTO="and SDDCTO like '%"+sSDDCTO+"%'"
      context['sSDDCTO'] =sSDDCTO
    sSDDOCO=request.GET['SDDOCO'].upper()
    if len(sSDDOCO)==0:
      qSDDOCO=''
    else:
      qSDDOCO="and SDDOCO like '%"+sSDDOCO+"%'"
      context['sSDDOCO'] =sSDDOCO
    com= request.GET['commit']
    if com =='更新發票日':
      
      ORADB = CONORACLE("SELECT * FROM (SELECT * FROM (select SDDCTO,SDDOCO,SDTRDJ,SDADDJ from F4211 WHERE (SDIVD<>SDTRDJ OR SDADDJ<>SDTRDJ) AND SDNXTR='620' AND SDTRDJ>='"+sd
                     +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDADDJ) WHERE SDADDJ <>0 "
                     +"union all SELECT * FROM("
                     +"select SDDCTO,SDDOCO,SDTRDJ,SDADDJ from F4211 WHERE SDIVD<>SDTRDJ  AND SDNXTR='620' AND SDTRDJ>='"+sd
                     +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDADDJ)  WHERE SDADDJ =0) "
                     +"  ORDER BY SDTRDJ DESC ,SDDOCO")
      for u in ORADB:
        ILTRDJ=str(u[2])
        ILDOCO=str(u[1])
        ILDCTO=str(u[0])
        #ILITM=str(u[4])
        if u[3]>0:
          ORADB = CONORACLE("update F4111 set ILTRDJ='"+ILTRDJ+"'  WHERE ILDOCO='"+ILDOCO+"' and ILDCTO='"+ILDCTO+"' "+'\n')
          ORADB = CONORACLE("update F4211 set SDIVD=SDTRDJ ,SDADDJ=SDTRDJ   WHERE SDDOCO='"+ILDOCO+"' and sddcto='"+ILDCTO+"'"+'\n')
        else:
          ORADB = CONORACLE("update F4211 set SDIVD=SDTRDJ  WHERE SDDOCO='"+ILDOCO+"' and sddcto='"+ILDCTO+"'"+'\n')
    #查詢:當沒有出貨日期時，訂單與發票日期不一致才要出現。
    '''
    f.write("SELECT * FROM (SELECT f42.*,f01.abalph FROM (select SDDCTO,SDDOCO,sdpa8,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDTRDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDTRDJ,4,3)-1,'yyyymmdd'))) AS SDTRDJ"
	                   +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDIVD,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDIVD,4,3)-1,'yyyymmdd'))) AS SDIVD"
					   +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDADDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDADDJ,4,3)-1,'yyyymmdd'))) AS SDADDJ"
					   +" from F4211 WHERE (SDIVD<>SDTRDJ OR SDADDJ<>SDTRDJ) AND SDNXTR='620' AND SDTRDJ>='"+sd
                       +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDIVD,SDADDJ,sdpa8)f42,f0101 f01  WHERE SDADDJ IS NOT NULL and f01.aban8=f42.sdpa8"
                       +" UNION ALL SELECT f42.*,f01.abalph FROM("
					   +"select SDDCTO,SDDOCO,sdpa8,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDTRDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDTRDJ,4,3)-1,'yyyymmdd'))) AS SDTRDJ"
                       +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDIVD,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDIVD,4,3)-1,'yyyymmdd'))) AS SDIVD"
                       +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDADDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDADDJ,4,3)-1,'yyyymmdd'))) AS SDADDJ"
                       +" from F4211 WHERE SDIVD<>SDTRDJ  AND SDNXTR='620' AND SDTRDJ>='"+sd
                       +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDIVD,SDADDJ,sdpa8)f42,f0101 f01   WHERE SDADDJ IS NULL and f01.aban8=f42.sdpa8)"
                       +"  ORDER BY SDTRDJ DESC ,SDDOCO")    
    '''
    ORADB = CONORACLE("SELECT * FROM (SELECT f42.*,f01.abalph FROM (select SDDCTO,SDDOCO,sdpa8,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDTRDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDTRDJ,4,3)-1,'yyyymmdd'))) AS SDTRDJ"
	                   +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDIVD,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDIVD,4,3)-1,'yyyymmdd'))) AS SDIVD"
					   +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDADDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDADDJ,4,3)-1,'yyyymmdd'))) AS SDADDJ"
					   +" from F4211 WHERE (SDIVD<>SDTRDJ OR SDADDJ<>SDTRDJ) AND SDNXTR='620' AND SDTRDJ>='"+sd
                       +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDIVD,SDADDJ,sdpa8)f42,f0101 f01  WHERE SDADDJ IS NOT NULL and f01.aban8=f42.sdpa8"
                       +" UNION ALL SELECT f42.*,f01.abalph FROM("
					   +"select SDDCTO,SDDOCO,sdpa8,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDTRDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDTRDJ,4,3)-1,'yyyymmdd'))) AS SDTRDJ"
                       +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDIVD,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDIVD,4,3)-1,'yyyymmdd'))) AS SDIVD"
                       +",TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(SDADDJ,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(SDADDJ,4,3)-1,'yyyymmdd'))) AS SDADDJ"
                       +" from F4211 WHERE SDIVD<>SDTRDJ  AND SDNXTR='620' AND SDTRDJ>='"+sd
                       +"' and SDTRDJ<='"+ed+"' "+stype+qSDDCTO+qSDDOCO+" group by SDDCTO,SDDOCO,SDTRDJ,SDIVD,SDADDJ,sdpa8)f42,f0101 f01   WHERE SDADDJ IS NULL and f01.aban8=f42.sdpa8)"
                       +"  ORDER BY SDTRDJ DESC ,SDDOCO")	
    
    for o in ORADB:
      TF4211=[]
      TF4211.append(str(o[0]))      
      TF4211.append(str(o[1]))
      TF4211.append(str(o[6]))
      TF4211.append(str(o[3]))
      TF4211.append(str(o[4]))
      TF4211.append(str(o[5]))
      #TF4211.append(str(o[5]))
      #TF4211.append(str(o[6]))
      #TF4211.append(str(o[7]))
      F4211.append(TF4211)	  
    context['F4211']=F4211
    if len(F4211)==0:
      context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(F4211))+" 筆資料"
    context['reportmes']='<B>森邦(股) 銷貨F4211日期檢查明細('+sday+'~'+eday+')</B>'	
  except:
    s=''    
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
    context['F4211']=F4211
    context['CK2'] = 'on'
  #f.close()
  return render(request, 'F4211.html',context )#傳入參數
def F4211saleitem(request):
  context= {}
  #f=open(r'C:\Users\chris\chrisdjango\error.txt','w')
  F4211_1=[]
  try:    
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    context['Sday'] = sday
    context['Eday'] = eday
    #f.write(eday+'\n')	
    stype=''
    ST=[]
    sdan8=request.GET['an8']
    context['san8'] = sdan8
    #f.write(sdan8+'\n')
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    sSDDCTO=request.GET['SDDCTO'].upper()
    if len(sSDDCTO)==0:
      qSDDCTO=''
    else:
      qSDDCTO="and SDDCTO like '%"+sSDDCTO+"%'"
      context['sSDDCTO'] =sSDDCTO
    sSDDOCO=request.GET['SDDOCO'].upper()
    if len(sSDDOCO)==0:
      qSDDOCO=''
    else:
      qSDDOCO="and SDDOCO like '%"+sSDDOCO+"%'"
      context['sSDDOCO'] =sSDDOCO
    #com= request.GET['commit']
    
    					 
    ORADB = CONORACLE("SELECT f42.*,f01.abalph FROM (select SDDCTO,SDDOCO,sdpa8 from F4211 WHERE  SDTRDJ>='"+sd  #SDNXTR='620' AND
                       +"' and SDTRDJ<='"+ed+"' "+qSDDCTO+qSDDOCO+" and sdpa8 like '"+sdan8+"%' group by SDDCTO,SDDOCO,sdpa8)f42,f0101 f01  WHERE f01.aban8=f42.sdpa8 order by f42.SDDOCO")  
    					   
    
    for o in ORADB:
      TF4211=[]
      TF4211.append(str(o[0]))      
      TF4211.append(str(o[1]))
      TF4211.append(str(o[3]))
      TF4211.append(str(o[2]))
      #TF4211.append(str(o[4]))
      #TF4211.append(str(o[5]))
      #TF4211.append(str(o[5]))
      #TF4211.append(str(o[6]))
      #TF4211.append(str(o[7]))
      F4211_1.append(TF4211)	  
    context['F4211_1']=F4211_1
    if len(F4211_1)==0:
      context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(F4211_1))+" 筆資料"
    context['reportmes']='<B>森邦(股) 銷貨明細('+sday+'~'+eday+')</B>'	
  except:
    s=''    
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday
    context['Eday'] = nday
    context['F4211_1']=F4211_1
  #f.close()
  return render(request, 'F4211saleitem.html',context )#傳入參數
def ReceivableItem(request):
  context= {}
  #f=open(r'C:\Users\chris\chrisdjango\error.txt','w')
  ReItem=[]
  try:    
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    context['Sday'] = sday
    context['Eday'] = eday
    #f.write(eday+'\n')	
    stype=''
    ST=[]
    sdan8=request.GET['custno']
    #f.write(sdan8+'\n')
    context['custno'] = sdan8
    sd=(sday[:4]+sday[5:7]+sday[8:10])
    ed=(eday[:4]+eday[5:7]+eday[8:10])
    #f.write(ed+'\n')
    '''
    sSDDCTO=request.GET['SDDCTO'].upper()
    if len(sSDDCTO)==0:
      qSDDCTO=''
    else:
      qSDDCTO="and SDDCTO like '%"+sSDDCTO+"%'"
      context['sSDDCTO'] =sSDDCTO
    sSDDOCO=request.GET['SDDOCO'].upper()
    if len(sSDDOCO)==0:
      qSDDOCO=''
    else:
      qSDDOCO="and SDDOCO like '%"+sSDDOCO+"%'"
      context['sSDDOCO'] =sSDDOCO
    #com= request.GET['commit']
    '''
    title=['<th style="width:5%;">帳款日期</th>','<th style="width:10%;">地址號</th>','<th style="width:15%;">名稱</th>','<th style="width:10%;">單據憑證</th>','<th style="width:10%;">發票號碼</th>','<th style="width:10%;">應收總計</th>','<th style="width:10%;">已收金額</th>','<th style="width:10%;">本期應收</th>','<th style="width:5%;">備註</th>']	  
    ORADB = CONORACLE("select im.rpdgj,im.an8, im.abalph,  im.rpsdoc, im.rpdoc||'-'|| im.rpdct rpdoc,IM.TOTO1 as toto1,abs(B.TOTO1) as toto2, im.toto1+ b.toto1 as toto3 "
                     +" from VS_SALEACRE_RIRM  im left join VS_SALEACRE_RB  b on im.sddoco= b.sddoco  where  im.rpdgj >= '"+sd+"' and im.rpdgj <= '"+ed+"' and im.an8 like '"+sdan8
					 +"%' order by an8,rpdgj")
    					   
    
    for o in ORADB:
      TReItem=[]
      TReItem.append('<td style="width:5%;">'+str(o[0])+"</td>")      
      TReItem.append('<td style="width:10%;">'+str(o[1])+"</td>")
      TReItem.append('<td style="width:15%;">'+str(o[2])+"</td>")
      TReItem.append('<td style="width:10%;">'+str(o[3])+"</td>")
      TReItem.append('<td style="width:10%;">'+str(o[4])+"</td>")
      if str(o[5])=='None':
        TReItem.append('<td style="width:10%;">0</td>')
      else: TReItem.append('<td style="width:10%;">'+str(o[5])+"</td>")
      if str(o[6])=='None':
        TReItem.append('<td style="width:10%;">0</td>')
      else: TReItem.append('<td style="width:10%;">'+str(o[6])+"</td>")
      if str(o[7])=='None':
        TReItem.append('<td style="width:10%;">0</td>')
      else: TReItem.append('<td style="width:10%;">'+str(o[7])+"</td>")
      TReItem.append('<td style="width:5%;"></td>')
      ReItem.append(TReItem)	  
    context['ReItem']=ReItem
    if len(TReItem)==0:
      context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(ReItem))+" 筆資料"
    context['title']=title
    context['reportmes']='<B>森邦(股) 應收帳款明細('+sdan8+' : '+sday+'~'+eday+')</B>'	
  except:
    s=''    
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday
    context['Eday'] = nday
    #context['TReItem']=F4211_1
  #f.close()
  return render(request, 'ReceivableItem.html',context )#傳入參數
def F4311item(request):
  F4311=[]
  context= {}
  #f=open('C:\\Users\\Administrator\\Desktop\\txtlog\\F4311item.txt','w')
  try:
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    compno=request.GET['compno']	
    context['Sday'] = sday
    context['Eday'] = eday
    context['compno'] = compno
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])    
      	  
    title=['<th style="width:6%;">採購單號</th>','<th style="width:6%;">總帳日期</th>','<th style="width:15%;">廠商名稱</th>','<th style="width:5%;">應付總帳</th>','<th style="width:5%;">料號總帳</th>'
	   ,'<th style="width:5%;">料號</th>','<th style="width:13%;">商品名稱</th>','<th style="width:11%;">商品說明1</th>','<th style="width:6%;">商品說明2</th>','<th style="width:3%;">數量</th>'
	   ,'<th style="width:3%;">單價</th>','<th style="width:4%;">未結金額</th>','<th style="width:4%;">未稅金額</th>','<th style="width:4%;">稅金</th>','<th style="width:4%;">含稅金額</th>']
					  
					  
    serdata=CONORACLE("select f43.pddoco,TO_CHAR(TO_NUMBER(TO_CHAR(TO_DATE('20'||SUBSTR(f43.PDDGL,2,2)||'-01-01','yyyy-mm-dd')+SUBSTR(f43.PDDGL,4,3)-1,'yyyymmdd'))) "
                      +"AS PDDGL,f01.abalph,sum(f43.PDAEXP) as PDAEXP,sum(f43.PDAREC) as PDAREC,f43.pdlitm,f41.imsrtx,f43.PDDSC1,f43.PDDSC2,f43.pduorg/10000 as pduorg "
                      +",round(f43.pdprrc/10000,4) as price,(f43.pdaopn) as pdaopn ,f43.pdaopn EXP, case WHEN PDtxa1='VT' THEN round((f43.pdaopn)*0.05) ELSE 0 END tax , case WHEN PDtxa1='VT' THEN (round((f43.pdaopn)*0.05)+f43.pdaopn) ELSE f43.pdaopn END tax  from f4311 f43,f0101 f01,f4101 f41 where pddgl >='"+sd+"' and pddgl <= '"+ed+"'"
                      +"AND  F43.pdlttr=220 AND F43.PDDCTO='OD' and f43.pdan8 like '%"+compno+"%' and f01.aban8=f43.pdan8 and f41.imlitm=F43.pdlitm "
                      +"group by f43.pddoco,f01.abalph,f43.pdlitm,f41.imsrtx,f43.pduorg,f43.pdprrc,f43.pdaopn,f43.PDDGL ,f43.PDtxa1,f43.PDDSC1,f43.PDDSC2 ORDER BY pddoco desc")
    context['reportmes']='<B>森邦(股) 驗收進貨單('+sday+'~'+eday+')</B>'
    context['title'] = title
    
    #f.write(str(title))
    ds=0
    EXP_SUM = 0
    TAX_SUM = 0
    EXPTAX_SUM= 0
    for data in serdata:
      tf43=[]
      ds=ds+1
      #tf43.append('<td style="width:4%;">'+str(ds)+'</td>')
      tf43.append('<td style="width:5%;">'+str(data[0])+' '+'</td>')
      tf43.append('<td style="width:5%;">'+str(data[1])+'</td>')
      tf43.append('<td style="width:15%;">'+str(data[2])+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[3],',')+'</td>')
      tf43.append('<td style="width:5%;">'+str(data[4])+'</td>')
      tf43.append('<td style="width:5%;">'+str(data[5])+'</td>')
      tf43.append('<td style="width:10%;">'+str(data[6])+'</td>')
      tf43.append('<td style="width:10%;">'+str(data[7])+'</td>')
      tf43.append('<td style="width:10%;">'+str(data[8])+'</td>')
      tf43.append('<td style="width:5%;">'+str(data[9])+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[10],',')+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[11],',')+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[12],',')+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[13],',')+'</td>')
      tf43.append('<td style="width:5%;">'+format(data[14],',')+'</td>')

      EXP_SUM = EXP_SUM + int(data[12])
      TAX_SUM = TAX_SUM + int(data[13])
      EXPTAX_SUM = EXPTAX_SUM + int(data[14])
      #f.write(str(tf43))
      F4311.append(tf43)
    F4311.append(['<td style="width:5%;"></td>', '<td style="width:5%;"></td>', '<td style="width:15%;"></td>', '<td style="width:5%;"></td>', '<td style="width:5%;"></td>', '<td style="width:5%;"></td>', '<td style="width:10%;"></td>', '<td style="width:10%;"></td>', '<td style="width:10%;"></td>', '<td style="width:5%;"></td>', '<td style="width:5%;"></td>', '<td style="width:5%;">總計</td>', '<td style="width:5%;">'+format(EXP_SUM,',')+'</td>', '<td style="width:5%;">'+format(TAX_SUM,',')+'</td>', '<td style="width:5%;">'+format(EXPTAX_SUM,',')+'</td>'])
	
    context['F4311'] = F4311
    
    if str(len(F4311))=='0':
        context['mess']='查無資料'
    else:
      #f.write("共 "+str(len(F43121))+" 筆資料"+'\n')
      context['mess']="共 "+str(len(F4311)-1)+" 筆資料"		
  except:
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
  return render(request, 'F4311item.html',context )#傳入參數
  #f.write(str(context))
  #f.close()
def pccss(request):  
  return render(request, 'pc.css', )
def tablecss(request):  
  return render(request, 'table.css', )
def spcss(request):  
  return render(request, 'sp.css', )
def pettycash(request):#零用金待轉項目查詢
  # f = open(r'C:\Users\Administrator\Desktop\txtlog\pettycash.txt','w')
  context={}
  connection214=pyodbc.connect('DRIVER={SQL Server};SERVER=192.168.0.214;DATABASE=erps;UID=apuser;PWD=0920799339')
  cur214hd = connection214.cursor()
  try:
    sday=request.GET['sdate']#get values
    eday=request.GET['edate']
    context['Sday']=sday
    context['Eday']=eday
    L1=[]#MYSQL內EIP的list
    L2=[]#MSSQL內EIP的list
    aNo=[]#會科編號list
    gday=[]#總帳日期
    form=[]
    ser=CONMYSQL("SELECT h.formsflow_id FROM hplus_formsflow h where"+
    " formsflow_form IN('5431','5477') and formsflow_finished >= '"+sday+" 00:00:00' and formsflow_finished <= '"+eday+" 23:59:59' and formsflow_state>1")
    for s in ser:
      cun = str(s[0])
      cuns = cun.strip(' ')
      L1.append(cuns)
    cur214hd.execute("SELECT [eip_id] FROM [ERPS].[dbo].[eipfreturnlist]")
    for c in cur214hd:
      cun = str(c[0])
      cuns = cun.strip(' ')
      L2.append(cuns)
    s1=set(L1)
    s2=set(L2)
    L3=list(s1.difference(s2))#取出L1內不存在於L2的EIP編號並建立list
    for s in range(len(L3)):
      ser=CONMYSQL("SELECT h.formsflow_id,h.formsflow_version FROM hplus_formsflow h where h.formsflow_id = '"+str(L3[s])+"'")
      for s in ser:
        sereip=CONMYSQL("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
        " where formsflow_field_formsflow = '"+str(s[0])+"' and formsflow_field_version = '"+str(s[1])+"' and formsflow_field_name = '__paymethod'")
        for se in sereip:
          aNo.append(se[1])
    aNol=list(set(aNo))
    context['aNo']=aNol
    aNo=request.GET['aNo']
    L4=[]
    for s in range(len(L3)):
      ser=CONMYSQL("SELECT h.formsflow_id,h.formsflow_version FROM hplus_formsflow h where h.formsflow_id = '"+str(L3[s])+"'")
      for s in ser:
        sereip=CONMYSQL("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
        " where formsflow_field_formsflow = '"+str(s[0])+"' and formsflow_field_version = '"+str(s[1])+"' and formsflow_field_name = '__paymethod'")
        for sp in sereip:
          if sp[1]==aNo:
            sereipt=CONMYSQL("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
            " where formsflow_field_formsflow = '"+str(s[0])+"' and formsflow_field_version = '"+str(s[1])+"' and formsflow_field_name = '__gldate'")
            for st in sereipt:
              gday.append(st[1])
              L4.append(s[0])
    gdays=list(set(gday))
    context['gday']=gdays
    if request.method == "POST":#如果是POST才執行
      eipid = request.POST.getlist('eipid')
    #   EIPPAY2JDE(eipid)
      context['mess']='已轉入，請重新查詢,轉入總金額為' +str(EIPPAY2JDE(eipid))+'元'
    else:
      gday=request.GET['gday']
      for s in range(len(L4)):
        ser=CONMYSQL("SELECT h.formsflow_id,h.formsflow_version FROM hplus_formsflow h where h.formsflow_id = '"+str(L4[s])+"'")
        for s in ser:
          sereipt=CONMYSQL("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
          " where formsflow_field_formsflow = '"+str(s[0])+"' and formsflow_field_version = '"+str(s[1])+"' and formsflow_field_name = '__gldate'")
          for st in sereipt:
            if st[1]==gday:
              data=CONMYSQL("SELECT h.formsflow_id,h.formsflow_version,h.formsflow_display,h.formsflow_method,h.formsflow_finished,h.formsflow_passby FROM hplus_formsflow h"+
              " where h.formsflow_id='"+str(s[0])+"' and h.formsflow_version='"+str(s[1])+"'")
              for d in data:
                forms=[]
                forms.append(d[0])
                forms.append(d[2])
                form.append(forms)
      context['payform'] = form
      if len(form)==0:
        context['mess']='查無資料'
      else:
        context['mess']="共 "+str(len(form))+" 筆資料"
  except:
    s=''
  # f.close()
  return render(request, 'pettycash.html',context)
def paylist(request):#零用金待轉項目明細
  context={}
  # f = open(r'D:\chrisdjango\cashlist.txt','w')
  eid=request.GET['eipid']
  vs=CONMYSQL("SELECT h.formsflow_version,formsflow_owner FROM hplus_formsflow h where h.formsflow_id = '"+eid+"'")
  for v in vs:
    ver=v[0]
    id=v[1]
  eno=CONMYSQL("select empname,empno from hplus_special where account_id='"+str(id)+"'")
  for n in eno:
      context['name']=str(n[0]) + str(n[1])
  detail=CONMYSQL("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
  " where formsflow_field_formsflow = '"+eid+"' and formsflow_field_version = '"+str(ver)+"'")
  # f.write("select formsflow_field_name,formsflow_field_value FROM hplus_formsflow_field "+
  # " where formsflow_field_formsflow = '"+eid+"' and formsflow_field_version = '"+str(ver)+"'")
  for d in detail:
    if d[0]=='__tpa':
      context['tpay']=str(d[1])
    if d[0]=='__paydate':
      context['payday']=str(d[1])
    if d[0]=='__paymethod':
      context['paymethod']=str(d[1])
    if d[0]=='__date':
      context['date']=str(d[1])
    if d[0]=='__bankuser':
      context['bank']=str(d[1])
    if d[0]=='__chargefee':
      context['chargefee']=str(d[1])
    if d[0]=='__account':
      context['bankacc']=str(d[1])
    if d[0]=='__bank':
      context['bankname']=str(d[1])
    if d[0]=='__gldate':
      context['fpayday']=str(d[1])
  group=[]
  item=[]
  dep=[]
  ntp=[]
  t=[]
  tp=[]
  TAX=[]
  als=[]
  for dl in detail:
    if dl[0].find('__dep')==0:
      dep.append(dl[1])
      # f.write(str(dep))
    if dl[0].find('__group')==0:
      group.append(dl[1])
      # f.write(str(group))
    if dl[0].find('__item')==0:
      item.append(dl[1])
      # f.write(str(item))
    if dl[0].find('__ntp')==0:
      ntp.append(dl[1])
      # f.write(str(ntp))
    if dl[0].find('__t')==0:
      t.append(dl[1])
      # f.write(str(t))
    if dl[0].find('__TAX')==0:
      TAX.append(dl[1])
      # f.write(str(TAX))
    if dl[0].find('__tp')==0:
      tp.append(dl[1])
      # f.write(str(tp))
  for l in range(len(group)):
    all=[]
    if group[l]=='':
      break
    else:
      all.append(group[l])
      all.append(item[l])
      all.append(dep[l])
      all.append(ntp[l])
      all.append(t[l])
      all.append(tp[l])
      all.append(TAX[l])
      als.append(all)
  context['show'] = als
  return render(request, 'paylist.html',context)
def EIPPAY2JDE(eipid):#零用金轉JDE
  # OR_DATAID='CRPDTA'#測試區
  OR_DATAID='PRODDTA'#正式區
  #f = open(r'C:\Users\Administrator\Desktop\txtlog\EIPPAY2JDE.txt','w')
  f = open(r'C:\Users\Edward\Desktop\txtlog\EIPPAY2JDE.txt','w')
  

  VNEDUS=[]
  context={}
  connection214=pyodbc.connect('DRIVER={SQL Server};SERVER=192.168.0.214;DATABASE=erps;UID=apuser;PWD=0920799339')
  cur214hd = connection214.cursor()
  cur214hd.execute("SELECT max(rbatch) FROM [ERPS].[dbo].[eipfreturnlist]")

  for c in cur214hd:#做新的批次號
    VNEDBT=c[0]
  VNEDBT=int(VNEDBT)
  VNEDBT+=1
  for s in range(len(eipid)):
         
    ser=CONMYSQL("SELECT formsflow_passby FROM hplus_formsflow h where h.formsflow_id = '"+str(eipid[s])+"'")
    for e in ser:
      mans=str(e[0]).split(')(')
      man=mans[-1]
      man=man.strip(')')

      memid=CONMYSQL("SELECT empno FROM hplus_special where account_id  ='"+str(man)+"'")
      for xyz in memid:
        VNEDUS.append(xyz[0])
        
  Dtime = time.strftime("%H%M%S", time.localtime())
  c = 0 #項次計數器
  sumtpay = [] # 算出轉入總金額 
  for r in range(len(eipid)):
    #重複就不執行
    rep_xx = CONMSSQL214("SELECT count(*) FROM  [ERPS].[dbo].[eipfreturnlist] WHERE eip_id = '"+str(eipid[r])+"'")

    if str(rep_xx[0][0]) == '0':
      ser=CONMYSQL("SELECT h.formsflow_id,h.formsflow_version,right(h.formsflow_display,'7') display FROM hplus_formsflow h where h.formsflow_id = '"+str(eipid[r])+"'")
      for s in ser:   

        #data=CONMYSQL("SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'1',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%1%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'2',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%2%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'3',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%3%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'4',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%4%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'5',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%5%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'6',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%6%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'7',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%7%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'8',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%8%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%')")
        data=CONMYSQL("SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'1',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%1%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'2',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%2%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'3',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%3%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'4',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%4%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'5',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%5%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'6',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%6%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'7',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%7%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'8',''),''':''', replace(replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),'//',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%8%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%')")
        
        # f.write(str("SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'1',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%1%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'2',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%2%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'3',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%3%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'4',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%4%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'5',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%5%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'6',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%6%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'7',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%7%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%') UNION ALL SELECT CONCAT('{',GROUP_CONCAT('''',replace(formsflow_field_name,'8',''),''':''', replace(replace(if(formsflow_field_value ='' OR formsflow_field_value is null  ,' ',formsflow_field_value),',',''),'''',''),''''),'}') FROM hplus_formsflow_field A WHERE 1=1 AND formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND ((formsflow_field_name LIKE '__%8%' OR formsflow_field_name IN('__date','__gldate','__paydate','__paymethod','__tpa','__name'))AND formsflow_field_name NOT LIKE '__幣別%' AND formsflow_field_name NOT LIKE '__foreign%' AND formsflow_field_name NOT LIKE '__km%' AND formsflow_field_name NOT LIKE '__google%' AND formsflow_field_name NOT LIKE 'AR0%')"))
      
        VNEDDT =''
        VNDGJ  =''
        VNOBJ  =''
        NSUB   =''
        VNTORG =''
        Name   =''
        invITEM ={} # eipid 表單的發票 張數
        invITEM2 ={} #比對 同張發票最後一行是哪個
        TAXs = {} #稅金
        #VNEXR 擷取字串用
        VNEXRstr = str(s[2]) 
        for d in data:
          mstr = d[0]
          json= eval(mstr)
          #分支沒有空值(' ')才跑資料
          if json['__group'] != ' ':
            #行號
            c += 1
            i = c * 1000  
            #分支
            VNMCU = json['__group']
            #會科.子目
            item  = json['__item']
            items=str(item).split('.',1)
            try:#若分割的右邊值不存在，會進行例外執行
              #會科
              VNOBJ=str(items[0])
              #子目
              NSUB = str(items[1])      
            except:
              VNOBJ =str(item)
              NSUB = '  '
              #分支.會科.子目
            VNANI=VNMCU+'.'+item
            #現在日期
            cur214hd.execute("SELECT '1'+SUBSTRING(CONVERT(char(10), getdate(), 112),3,2)+RIGHT(REPLICATE('0','3') + CAST(datepart(dayofyear,CONVERT(char(10), getdate(), 112)) as NVARCHAR)  , 3) AS JDE")
            for yy in cur214hd:
              VNEDDT=yy[0]
            #申請日期
            # day   = json['__date'] 
            #申請原因  
            txt   = json['__dep']  
            #總帳日期 - 轉JDE格式
            glday = json['__gldate'] 
            cur214hd.execute("SELECT '1'+SUBSTRING('"+ str(glday) +"',3,2)+RIGHT(REPLICATE('0','3') + CAST(datepart(dayofyear,'" + str(glday) + "') as NVARCHAR)  , 3) AS JDE")
            for xx in cur214hd:
              VNDGJ = xx[0]
            # 使用者帳號_ID
            name = json['__name']
            serchname=CONMYSQL('SELECT empno,empname FROM hplus_special where account_id = "'+str(name)+'"')
            for serch in serchname:
              # 工號
              VNTORG = str(serch[0])
              # 姓名
              Name = str(serch[1])
            VNEXR = Name + txt
            #字串超過30截斷
            VNEXR=VNEXR[0:30]
            VNAA = json['__ntp']
            if VNAA == ' ':
              VNAA ='0'
            # json['__p']
            # json['__paydate']
            paymethod = json['__paymethod']
            # json['__sign']
            # json['__tp']
            tpa = json['__tpa']    
            if tpa == ' ':
              tpa ='0'  

            #f.write(
            #"INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
            #" VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
            #"VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
            #"'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"','"+VNANI+"','"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
            #",'"+str(NSUB)+"','"+str(VNAA)+"','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+str(VNTORG)+"','"+str(VNTORG)+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+str(VNAA)+"')"+'\n'
            #)
    
            CONORACLE(
            "INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
            " VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
            "VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
            "'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"','"+VNANI+"','"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
            ",'"+str(NSUB)+"','"+str(VNAA)+"','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+str(VNTORG)+"','"+str(VNTORG)+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+str(VNAA)+"')"+'\n'
            )

            invTAX = json['__TAX'] #發票
            TAXCASH = json['__t'] # 稅金
            #發票None不處理
            if invTAX != ' ': 
              #稅金加總
              getTAX = TAXs.get(invTAX)
              if getTAX != None:
                TaxSum = getTAX
                TaxTotal = int(TaxSum) + int(TAXCASH)
                del TAXs[invTAX]
                TAXs.setdefault(invTAX,TaxTotal)
              else:
                TAXs.setdefault(invTAX,TAXCASH)

              #判斷發票比對次數 
              getinvITEM2 = invITEM2.get(invTAX)
              if getinvITEM2 != None:
                #發票重複的次數
                invITEM2_COUNT = int(getinvITEM2)
                invITEM2_COUNT += 1
                #處理後刪除 KEY 後面再塞入值
                del invITEM2[invTAX]
                invITEM2.setdefault(invTAX,invITEM2_COUNT)
              else:
                #是None就塞1
                invITEM2.setdefault(invTAX,1)
              #查詢 同張發票出現的次數
              inv = CONMYSQL("SELECT formsflow_field_value, COUNT(*) FROM hplus_formsflow_field WHERE  formsflow_field_formsflow = '"+str(s[0])+"' AND formsflow_field_version = '"+str(s[1])+"' AND formsflow_field_name LIKE '__TAX%' AND formsflow_field_value > '' AND formsflow_field_value = '"+ invTAX +"' GROUP BY formsflow_field_value")
              for zz in inv:
                invITEM.setdefault(zz[0],zz[1])

              inv1 = invITEM.get(invTAX)  #同張單發票重複的次數
              inv2 = invITEM2.get(invTAX) #發票重複第幾次
              TAXTO = TAXs.get(invTAX)     #稅金總金額

              #筆數對應成功後塞入稅金
              if inv1 == inv2:
                
                # eipid分別增加稅金
                #行號
                c += 1
                i = c * 1000 
                  #會計科目固定
                VNANI = '      100000.1257        '
                VNMCU = '      100000'
                VNOBJ = '1257'
                VNSUB = ''
                  #稅金總數
                  
                TAX = TAXTO
                if TAX == ' ':
                  TAX = '0'
              
                  #  #說明-變更
                    
                VNEXR = invTAX
                VNEXR=VNEXR[0:30]
                  
                #f.write(
                #"INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
                #" VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
                #"VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
                #"'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"' ,'"+str(VNANI)+"' ,'"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
                #",'"+str(VNSUB)+"','0','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+VNEDUS[0]+"','"+VNEDUS[0]+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+ str(TAX)+"')"+'\n'
                #)
          
                CONORACLE(
                "INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
                " VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
                "VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
                "'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"' ,'"+str(VNANI)+"' ,'"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
                ",'"+str(VNSUB)+"','0','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+VNEDUS[0]+"','"+VNEDUS[0]+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+ str(TAX)+"')"+'\n'
                )
            
      #負項  eipid分別增加負項
      #行號
      
      c += 1
      i = c * 1000 
      #會科分割 如:   100000.1101.002  分成三等份 100000 = VNMCU,1101 =VNOBJ,002=VNSUB
      VNANI =  paymethod
      index1 = int(VNANI.find('.'))
      index2 = int(VNANI.find('.',index1+1))
        
      VNMCU = VNANI[0:6]
      VNOBJ = VNANI[index1+1:index2]
      #沒有子目
      if index2 != -1: 
        VNSUB = VNANI[index2+1:]
      else:
        VNSUB = ' '
        
      #總數變負數
      tpay = (int(tpa))*-1

      sumtpay.append(int(tpa)) #抓出每個表單的總金額

      #說明-變更
      VNEXR = "支付"+ VNEXRstr +str(Name)+"費用申請"  
      VNEXR=VNEXR[0:30]
        #---------------------------------------
        
      #f.write(
      #"INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
      #" VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
      #"VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
      #"'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"' ,'"+str(VNANI)+"' ,'"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
      #",'"+str(VNSUB)+"','0','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+VNEDUS[0]+"','"+VNEDUS[0]+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+ str(tpay)+"')"+'\n'
      #)
      CONORACLE(
      "INSERT INTO "+OR_DATAID+".F0911Z1(VNEDSQ,VNEDTN,VNEDER,VNEDDL,VNEDSP,VNEDTC,VNEDTR,VNEDAN,VNCO,VNAM,VNLT,VNPN,VNCTRY,VNCRCD,VNCRR,VNHCRR,VNODOC,"+
      " VNAN8,VNALTV,VNIVD,VNLNID,VNWY,VNWN,VNOPSQ,VNDOI,VNPID,VNJOBN,VNCRRM,VNEXR1,VNTXA1,VNTXITM,VNEDUS,VNEDLN,VNEDDT,VNEDBT,VNDGJ,VNANI,VNMCU,VNOBJ,VNSUB,VNAA,VNEXA,VNEXR,"+
      "VNTORG,VNUSER,VNUPMJ,VNUPMT,VNSTAM,VNAG) VALUES ('0','1','B','0','0','A','J','0','00100','2', 'AA','0','20','TWD','0','0','0','0','0','0','0','0','0','100','0',"+
      "'EP0911Z1 ','JDEWEB  ','D','  ','          ','0','"+VNEDUS[0]+"','"+str(i)+"','"+str(VNEDDT)+"','"+str(VNEDBT)+"','"+str(VNDGJ)+"' ,'"+str(VNANI)+"' ,'"+str(VNMCU)+"','"+str(VNOBJ)+"'"+
      ",'"+str(VNSUB)+"','0','EIP拋轉Oracle費用傳票','"+str(VNEXR)+"','"+VNEDUS[0]+"','"+VNEDUS[0]+"','"+str(VNEDDT)+"','"+str(Dtime)+"','0','"+ str(tpay)+"')"+'\n'
      )



      # ISERT 資料到 MSSQL 
      CONMSSQL214("INSERT INTO [ERPS].[dbo].[eipfreturnlist](eip_id,tdate,rbatch)VALUES('"+str(eipid[r])+"',CONVERT(varchar,GETDATE(),112),'"+ str(VNEDBT) +"')")
    else:
      context['mess']='重複點選按鈕'
  sumtpay1 = str(sum(sumtpay))
  return sumtpay1
  f.close()
def saleTotal(request): #客戶銷售總表
#   f = open(r'C:\Users\Administrator\Desktop\txtlog\saleTotal.txt','w')
  F43121=[]
  context= {}
  try:
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']
    compno=request.GET['compno']	
    context['Sday'] = sday
    context['Eday'] = eday
    context['compno'] = compno
    sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    try:
      ck1=request.GET['Checkbox1']
      context['ck'] ='ck1'
    except:
      ck1='off'
    try:
      ck2=request.GET['Checkbox2']
      context['ck'] ='ck2'
    except:
      ck2='off'
    if ck1=='on':

      title=['<th style="width:4%;">名次</th>','<th style="width:5%;">地址號</th>','<th style="width:5%;">廠商代碼</th>','<th style="width:12%;">名稱</th>','<th style="width:5%;">未稅金額</th>','<th style="width:5%;">稅金</th>','<th style="width:5%;">含稅金額</th>']

      #serdata=CONORACLE("SELECT GLAN8,ABALKY,ABALPH,(SUM(glaa)*-1) AMT, (SUM(glaa*0.05)*-1) TAX, (SUM(glaa)*-1)+(SUM(glaa*0.05)*-1) FROM F0911 A LEFT JOIN F0101 B ON B.ABAN8 = A.GLAN8 WHERE GLDGJ  BETWEEN "+sd+" AND "+ed+" AND GLDCT IN('RI','RM') AND A.GLANI NOT IN('      100000.5101            ','      100000.5201            ','      100000.1201            ','      100000.1202            ') AND TRIM(GLAN8) like '%"+compno+"%' AND GLLT ='AA' GROUP BY GLAN8,ABALKY,ABALPH ORDER BY GLAN8 ASC")
      serdata=CONORACLE("SELECT RPAN8,ABALKY,TO_CHAR(rpalph) rpalph,SUM(RPATXA) RPATXA,SUM(RPSTAM) RPSTAM,SUM(RPAG) RPAG FROM f03b11 A LEFT JOIN F0101 B ON B.ABAN8 = A.RPAN8 WHERE RPDGJ BETWEEN '"+sd+"' AND '"+ed+"' and RPDCT in('RI','RM') AND TRIM(RPAN8) like '%"+compno+"%' GROUP BY RPAN8,rpalph,ABALKY ORDER BY RPAN8 ASC")
      context['reportmes']='<B>森邦(股) 銷貨總金額排名('+sday+'~'+eday+')</B>'
      context['title'] = title
      ds=0
      for data in serdata:
        tf43=[]
        ds=ds+1 
       	
        tf43.append('<td style="width:4%;">'+str(ds)+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[0])+' '+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:12%;">'+str(data[2])+'</td>')
        tf43.append('<td style="width:5%;">'+format(data[3],',')+'</td>')#未稅金額
        tf43.append('<td style="width:5%;">'+format(data[4],',')+'</td>')#稅金
        tf43.append('<td style="width:5%;">'+format(data[5],',')+'</td>')#含稅金額
        F43121.append(tf43)
      
      context['saleTotal'] = F43121
      
    if ck2=='on':
      title=['<th style="width:5%;">地址號碼</th>','<th style="width:5%;">訂單號碼</th>','<th style="width:5%;">行號</th>','<th style="width:5%;">第二料號</th>','<th style="width:12%;">商品名稱</th>','<th style="width:5%;">未稅金額</th>','<th style="width:5%;">數量</th>']#,'<th style="width:20%;">廠商名稱</th>'

      serdata=CONORACLE("SELECT RPAN8,SDDCTO||'-'||SDDOCO DOCO,SDLNID/1000 LNID,SDLITM,SDDSC1,RPATXA,RPU FROM f03b11 A LEFT JOIN F4211 B ON B. SDDCTO = A.RPSDCT AND B.SDDOCO =A.RPSDOC AND B.SDLNID = A.RPLNID WHERE RPDGJ BETWEEN '"+sd+"' AND '"+ed+"' and RPDCT in('RI','RM') AND RPAN8 like '%"+compno+"%' ORDER BY SDDOCO,SDLNID ASC")
     
      context['title'] = title
      context['reportmes']='<B>森邦(股) 銷貨明細('+sday+'~'+eday+')</B>'
      ds=0
      for data in serdata:
        tf43=[]
        #ds=ds+1
    
        tf43.append('<td style="width:5%;">'+str(data[0])+'</td>')        
        tf43.append('<td style="width:5%;">'+str(data[1])+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[2])+' '+'</td>')
        tf43.append('<td style="width:5%;">'+str(data[3])+'</td>')
        tf43.append('<td style="width:12%;">'+str(data[4])+'</td>')
        tf43.append('<td style="width:5%;">'+format(data[5],',')+'</td>')#未稅金額
        tf43.append('<td style="width:5%;">'+format(data[6],',')+'</td>')#數量
		
        F43121.append(tf43)
      context['saleTotal'] = F43121	
    if str(len(F43121))=='0':
        context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(F43121))+" 筆資料"		
  except:
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
  return render(request, 'saleTotal.html',context )#傳入參數
  #f.write(str(context))
#   f.close()   
def invoiceCheck(request):
#   f = open(r'C:\Users\Administrator\Desktop\txtlog\invoiceCheck.txt','w')
  INV=[]
  context= {}
  Void = 0 #作廢張數
  TAXIDtotal = [] # 有統編金額
  TAXtotal = [] # 有統編稅額
  NOTAXIDtotal =[] # 無統編金額
  try:
    # f.write(str('OK')+'\n')
    sday=request.GET['Sday']#get values
    eday=request.GET['Eday']	
    context['Sday'] = sday
    context['Eday'] = eday
    # sd=getodate(sday[:4]+sday[5:7]+sday[8:10])
    # ed=getodate(eday[:4]+eday[5:7]+eday[8:10])
    
    # f.write(str(sday)+'\n')


    wb = Workbook()	
    ws1 = wb.active
    ws1.title = "銷貨總金額"
    ws1.append(['發票別','所屬年月','發票聯式','第冊數','發票號碼','統一編號','銷售額','稅別','稅額','月日','買受人','地址號'])

  
	
      #發票別,所屬年月,發票聯式,第冊數,發票號碼,統一編號,銷售額,稅別,稅額,月日,買受人,地址號
    title=['<th style="width:2%;">發票別</th>','<th style="width:3%;">所屬年月</th>','<th style="width:2%;">發票聯式</th>','<th style="width:3%;">第冊數</th>','<th style="width:5%;">發票號碼</th>','<th style="width:5%;">統一編號</th>','<th style="width:5%;">銷售額</th>','<th style="width:2%;">稅別</th>','<th style="width:5%;">稅額</th>','<th style="width:3%;">月日</th>','<th style="width:15%;">買受人</th>','<th style="width:5%;">地址號</th>']
    serdata=CONORACLE("SELECT FC 發票別, SUBSTR(CDATE-19110000,0,5) 所屬年月, '4 '發票聯式, '0000'第冊數, S_NO 發票號碼, CASE WHEN TOTO1=0 THEN '+' ELSE CUSTID END 統一編號,  CASE WHEN  CUSTID ='        '  THEN TO_NUMBER(TOTO1)+TO_NUMBER(TAM) ELSE TO_NUMBER(TOTO1) END  銷售額, '5'稅別, CASE WHEN  CUSTID ='        '  THEN '0' ELSE TAM END  稅額,  to_number(SUBSTR(CDATE,1,4))||'/'||to_number(SUBSTR(CDATE,5,2))||'/'||to_number(SUBSTR(CDATE,7,2)) 月日, ALPH 買受人, AN8 地址號 FROM proddta.VS_INV A WHERE CDATE  BETWEEN replace('"+sday+"','-','') AND replace('"+eday+"','-','')  AND FC LIKE '3%'")
    context['reportmes']='<B>每月發票('+sday+'~'+eday+')</B>'
    context['title'] = title
    # ds=0
    i=2
    for data in serdata:
      
      dttm = datetime.datetime.strptime(data[9], "%Y/%m/%d")
	  
      tf=[]
      tf.append(data[0]) 
      tf.append(data[1])     
      tf.append(data[2])     
      tf.append(data[3])
      tf.append(data[4])
      tf.append(data[5])
      tf.append(int(data[6]))
      tf.append(data[7])
      tf.append(int(data[8]))
      tf.append(dttm)
      tf.append(data[10])
      tf.append(data[11])
      INV.append(tf) 
      ws1.append(tf)
	  	  
      ws1.cell(row=i, column=10).number_format = 'm/d' #EXCEL 樣式 月日
        
      i+=1
	  
      if data[5].find('+')==0:
        Void += 1
    
      if data[5] != '        ':
        TAXIDtotal.append(int(data[6]))
        TAXtotal.append(int(data[8]))
      else:
        NOTAXIDtotal.append(int(data[6])+int(data[8]))

    context['invoiceCheck'] = INV
    
    if str(len(INV))=='0':
        context['mess']='查無資料'
    else:
      context['mess']="共 "+str(len(INV))+" 筆資料"		
      context['DataRow'] = "總計張數:"+str(len(INV))
    context['Void'] = "作廢張數:"+str(Void)
    context['TAXDATA1'] = '有統編:'+str(sum(TAXIDtotal))
    context['TAXDATA2'] = '稅額:'+str(sum(TAXtotal))
    context['TAXDATA3'] = '無統編:'+str(sum(NOTAXIDtotal))

    wb.save('C:\\Users\\Administrator\\chrisdjango\\MEDIA\\'+sday+'~'+eday+'發票查詢.xlsx')#正式路徑
    context['efilename']= sday+'~'+eday+'發票查詢.xlsx'
  except:
    # f.write(str('NO')+'\n')
    nday=showday(0,'-',0) #今天日期
    context['Sday'] = nday[:8]+'01'
    context['Eday'] = nday
  return render(request, 'invoiceCheck.html',context )#傳入參數
  # f.close() 
